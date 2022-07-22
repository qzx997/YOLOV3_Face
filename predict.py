# -----------------------------------------------------------------------#
#   预测
# -----------------------------------------------------------------------#
import time

import cv2
import numpy as np
import pandas as pd
from PIL import Image
import os

from yolo import YOLO
from openpyxl import load_workbook


def excel_write(cnt, row, img_name):
    wb = load_workbook(r'demo.xlsx')
    print(type(wb))
    sheet = wb.active
    sheet.cell(row=row, column=1).value = img_name
    for i in range(4):
        sheet.cell(row=row, column=i + 2).value = cnt[i]

    wb.save(r'demo.xlsx')


if __name__ == "__main__":
    yolo = YOLO()

    video_path = 0
    video_save_path = ""
    video_fps = 25.0

    test_interval = 100
    dir_origin_path = "img/"
    dir_save_path = "img_out/"


    img_path = 'VOCdevkit/predict/test_images/'
    imgs = os.listdir(img_path)
    # row = 1
    # num = []
    for img_name in imgs:
        img = img_path + img_name
        image = Image.open(img)
        r_image, cnt = yolo.detect_image(image)
        # print(cnt)
        # num.append(cnt)
        r_image.save('VOCdevkit/predict/test_outcome/' + img_name)
        # excel_write(cnt, row, img_name)

        # row += 1
    print(num)
    print(imgs)
    csv_save_data = pd.DataFrame(data=num)
    csv_save_label = pd.DataFrame(data=imgs)
    csv_save_data.to_csv('data_save.csv', encoding='gbk')
    csv_save_label.to_csv('data_save_label.csv', encoding='gbk')
